"""
Ruta API: Importación de datos desde JSON o Excel.

Soporta la estructura del Excel de panadería:
  - catalogo_ingredientes: lista de ingredientes con peso, costo, unidad
  - recetas: lista de recetas con ingredientes y resumen

Corrige automáticamente el error del Excel donde los ingredientes en
unidad G y UND tienen costo_por_gramo 1000x menor al correcto.
"""

import json
import io
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from app.database import get_db
from app.models.ingrediente import Ingrediente
from app.models.receta import Receta, RecetaDetalle
from app.models.producto import Producto
from app.models.movimiento import MovimientoStock
from app.services.ingrediente_service import calcular_costo_por_gramo, FACTORES_A_GRAMOS

router = APIRouter(prefix="/importar", tags=["Importación"])

# Mapa para normalizar unidades del Excel al sistema
MAPA_UNIDADES = {
    "KG": "kg", "G": "g", "GR": "g", "GRAMO": "g", "GRAMOS": "g",
    "LTS": "lts", "LT": "lts", "LITRO": "lts", "LITROS": "lts",
    "ML": "ml", "UND": "und", "UN": "und", "UNIDAD": "und", "UNIDADES": "und",
    "HOJ": "hoj", "HOJA": "hoj", "HOJAS": "hoj",
}


def normalizar_unidad(unidad: str) -> str:
    return MAPA_UNIDADES.get(unidad.upper().strip(), "und")


def calcular_costo_unitario_correcto(costo: float, peso: float, unidad: str) -> float:
    """
    Calcula el costo unitario correcto a partir del total y el peso/cantidad.

    En el Excel original, el costo_por_gramo está mal calculado para G y UND
    (se divide por 1000 extra). Aquí calculamos directamente desde la fuente:
      costo_unitario = costo_total / cantidad_en_unidad_base
    """
    if peso <= 0:
        return 0
    return costo / peso


async def _importar_ingredientes(db: AsyncSession, catalogo: list) -> dict:
    """Procesa el catálogo de ingredientes del JSON/Excel."""
    creados = []
    actualizados = []
    errores = []

    for item in catalogo:
        try:
            nombre = str(item.get("ingrediente", item.get("nombre", ""))).strip().lower()
            if not nombre:
                continue

            unidad_raw = str(item.get("unidad_medida", "und"))
            unidad = normalizar_unidad(unidad_raw)

            peso = float(item.get("peso", 0))
            costo = float(item.get("costo", 0))

            if peso <= 0 or costo <= 0:
                errores.append(f"{nombre}: peso o costo inválido (peso={peso}, costo={costo})")
                continue

            # Calcular costo_unitario correcto desde datos fuente
            costo_unitario = calcular_costo_unitario_correcto(costo, peso, unidad)
            costo_gramo = calcular_costo_por_gramo(costo_unitario, unidad)

            # Buscar si ya existe
            result = await db.execute(
                select(Ingrediente).where(Ingrediente.nombre == nombre)
            )
            existente = result.scalar_one_or_none()

            if existente:
                existente.costo_unitario = round(costo_unitario, 4)
                existente.costo_por_gramo = round(costo_gramo, 6)
                actualizados.append({
                    "nombre": nombre,
                    "unidad": unidad,
                    "costo_unitario": round(costo_unitario, 4),
                    "costo_por_gramo": round(costo_gramo, 6),
                })
            else:
                nuevo = Ingrediente(
                    nombre=nombre,
                    unidad_medida=unidad,
                    stock_actual=0,
                    stock_minimo=0,
                    costo_unitario=round(costo_unitario, 4),
                    costo_por_gramo=round(costo_gramo, 6),
                )
                db.add(nuevo)
                await db.flush()
                creados.append({
                    "nombre": nombre,
                    "unidad": unidad,
                    "costo_unitario": round(costo_unitario, 4),
                    "costo_por_gramo": round(costo_gramo, 6),
                })

        except Exception as e:
            errores.append(f"{item}: {str(e)}")

    return {"creados": creados, "actualizados": actualizados, "errores": errores}


async def _importar_recetas(db: AsyncSession, recetas_data: list) -> dict:
    """Procesa las recetas del JSON/Excel."""
    creadas = []
    omitidas = []
    errores = []

    for receta_data in recetas_data:
        try:
            nombre_receta = str(receta_data.get("nombre", "")).strip()
            if not nombre_receta:
                continue

            resumen = receta_data.get("resumen", {})
            rendimiento = int(resumen.get("unidades_producidas", 1) or 1)
            mano_de_obra = float(resumen.get("mano_de_obra", 0) or 0)
            ganancia = float(resumen.get("ganancia_porcentaje", 0) or 0)

            ingredientes_receta = receta_data.get("ingredientes", [])
            # Filtrar ingredientes con cantidad > 0
            ingredientes_validos = [
                i for i in ingredientes_receta
                if float(i.get("cantidad", 0)) > 0
            ]

            if not ingredientes_validos:
                omitidas.append(f"{nombre_receta}: sin ingredientes con cantidad > 0")
                continue

            # Buscar o crear el producto asociado
            nombre_producto = nombre_receta.lower()
            result = await db.execute(
                select(Producto).where(Producto.nombre == nombre_producto)
            )
            producto = result.scalar_one_or_none()

            if not producto:
                pvp = float(resumen.get("precio_venta_pvp", 0) or 0)
                producto = Producto(
                    nombre=nombre_producto,
                    precio_venta=round(pvp, 2),
                    stock_actual=0,
                    stock_minimo=0,
                    costo_unitario=0,
                )
                db.add(producto)
                await db.flush()

            # Verificar si ya existe la receta
            result2 = await db.execute(
                select(Receta).where(Receta.nombre == nombre_receta)
            )
            receta_existente = result2.scalar_one_or_none()
            if receta_existente:
                omitidas.append(f"{nombre_receta}: ya existe, se omite")
                continue

            # Crear la receta
            receta = Receta(
                producto_id=producto.id,
                nombre=nombre_receta,
                rendimiento=rendimiento,
                mano_de_obra=round(mano_de_obra, 4),
                porcentaje_ganancia=round(ganancia * 100 if ganancia < 1 else ganancia, 2),
                porcentaje_merma=0,
                notas=None,
            )
            db.add(receta)
            await db.flush()

            # Crear detalles (ingredientes de la receta)
            detalles_ok = []
            detalles_error = []
            for ing_data in ingredientes_validos:
                nombre_ing = str(ing_data.get("nombre", "")).strip().lower()
                cantidad = float(ing_data.get("cantidad", 0))
                unidad_raw = str(ing_data.get("unidad", "g"))
                unidad = normalizar_unidad(unidad_raw)

                # Buscar el ingrediente en la BD
                result3 = await db.execute(
                    select(Ingrediente).where(Ingrediente.nombre == nombre_ing)
                )
                ingrediente = result3.scalar_one_or_none()

                if not ingrediente:
                    detalles_error.append(f"ingrediente '{nombre_ing}' no encontrado")
                    continue

                detalle = RecetaDetalle(
                    receta_id=receta.id,
                    ingrediente_id=ingrediente.id,
                    cantidad=cantidad,
                    unidad=unidad,
                )
                db.add(detalle)
                detalles_ok.append(nombre_ing)

            creadas.append({
                "nombre": nombre_receta,
                "rendimiento": rendimiento,
                "ingredientes_ok": len(detalles_ok),
                "ingredientes_error": detalles_error,
            })

        except Exception as e:
            errores.append(f"{receta_data.get('nombre', '?')}: {str(e)}")

    return {"creadas": creadas, "omitidas": omitidas, "errores": errores}


def _parsear_json(contenido: bytes) -> dict:
    """Parsea el JSON de la estructura de panadería."""
    data = json.loads(contenido)
    return {
        "catalogo_ingredientes": data.get("catalogo_ingredientes", []),
        "recetas": data.get("recetas", []),
    }


def _parsear_excel(contenido: bytes) -> dict:
    """
    Parsea un archivo Excel con la estructura de panadería.

    Formato esperado del Excel:
    - Hoja 'Catalogo' o primera hoja: columnas ingrediente, unidad_medida, peso, costo
    - Hoja 'Recetas': columnas nombre (encabezado de receta), ingrediente, cantidad, unidad, costo
    """
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(contenido), data_only=True)
    catalogo = []
    recetas = []

    # Intentar leer hoja de catálogo
    hoja_catalogo = None
    for nombre_hoja in ["Catalogo", "catalogo", "Ingredientes", "ingredientes", "Catálogo"]:
        if nombre_hoja in wb.sheetnames:
            hoja_catalogo = wb[nombre_hoja]
            break
    if hoja_catalogo is None:
        hoja_catalogo = wb.worksheets[0]

    # Leer cabeceras de la primera fila
    headers = []
    for row in hoja_catalogo.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = [str(c).strip().lower() if c else "" for c in row]
        break

    # Mapear columnas
    col_map = {}
    for i, h in enumerate(headers):
        if any(k in h for k in ["ingrediente", "nombre", "material"]):
            col_map["ingrediente"] = i
        elif any(k in h for k in ["unidad", "medida"]):
            col_map["unidad_medida"] = i
        elif "peso" in h or "cantidad" in h or "cant" in h:
            col_map["peso"] = i
        elif "costo" in h or "precio" in h or "valor" in h:
            col_map["costo"] = i

    for row in hoja_catalogo.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        try:
            item = {
                "ingrediente": row[col_map["ingrediente"]] if "ingrediente" in col_map else "",
                "unidad_medida": row[col_map["unidad_medida"]] if "unidad_medida" in col_map else "und",
                "peso": float(row[col_map["peso"]] or 0) if "peso" in col_map else 0,
                "costo": float(row[col_map["costo"]] or 0) if "costo" in col_map else 0,
            }
            if item["ingrediente"] and item["peso"] > 0 and item["costo"] > 0:
                catalogo.append(item)
        except Exception:
            continue

    return {"catalogo_ingredientes": catalogo, "recetas": recetas}


@router.post("/preview")
async def preview_importacion(archivo: UploadFile = File(...)):
    """
    Muestra lo que se importaría sin ejecutar cambios en la BD.
    Soporta .json y .xlsx
    """
    contenido = await archivo.read()
    nombre = archivo.filename or ""

    if nombre.endswith(".json") or archivo.content_type == "application/json":
        data = _parsear_json(contenido)
    elif nombre.endswith((".xlsx", ".xls")):
        data = _parsear_excel(contenido)
    else:
        # Intentar como JSON por defecto
        try:
            data = _parsear_json(contenido)
        except Exception:
            raise HTTPException(400, "Formato no soportado. Use .json o .xlsx")

    # Preparar preview (sin tocar la BD)
    ingredientes_preview = []
    for item in data["catalogo_ingredientes"]:
        nombre_ing = str(item.get("ingrediente", "")).strip().lower()
        unidad = normalizar_unidad(str(item.get("unidad_medida", "und")))
        peso = float(item.get("peso", 0))
        costo = float(item.get("costo", 0))
        if not nombre_ing or peso <= 0 or costo <= 0:
            continue
        costo_unit = calcular_costo_unitario_correcto(costo, peso, unidad)
        costo_gramo = calcular_costo_por_gramo(costo_unit, unidad)
        ingredientes_preview.append({
            "nombre": nombre_ing,
            "unidad": unidad,
            "peso_compra": peso,
            "costo_compra": costo,
            "costo_unitario_calculado": round(costo_unit, 4),
            "costo_por_gramo_calculado": round(costo_gramo, 6),
        })

    recetas_preview = []
    for r in data["recetas"]:
        resumen = r.get("resumen", {})
        recetas_preview.append({
            "nombre": r.get("nombre", ""),
            "rendimiento": int(resumen.get("unidades_producidas", 1) or 1),
            "num_ingredientes": len([i for i in r.get("ingredientes", []) if float(i.get("cantidad", 0)) > 0]),
            "mano_de_obra": float(resumen.get("mano_de_obra", 0) or 0),
            "pvp_sugerido": float(resumen.get("precio_venta_pvp", 0) or 0),
        })

    return {
        "ingredientes": ingredientes_preview,
        "recetas": recetas_preview,
        "total_ingredientes": len(ingredientes_preview),
        "total_recetas": len(recetas_preview),
    }


@router.post("/ejecutar")
async def ejecutar_importacion(
    archivo: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
):
    """
    Importa datos desde un archivo JSON o Excel a la base de datos.
    Crea o actualiza ingredientes y recetas.
    """
    contenido = await archivo.read()
    nombre = archivo.filename or ""

    if nombre.endswith(".json") or archivo.content_type == "application/json":
        data = _parsear_json(contenido)
    elif nombre.endswith((".xlsx", ".xls")):
        data = _parsear_excel(contenido)
    else:
        try:
            data = _parsear_json(contenido)
        except Exception:
            raise HTTPException(400, "Formato no soportado. Use .json o .xlsx")

    # 1. Importar ingredientes primero
    resultado_ingredientes = await _importar_ingredientes(db, data["catalogo_ingredientes"])

    # 2. Importar recetas (dependen de ingredientes ya creados)
    resultado_recetas = await _importar_recetas(db, data["recetas"])

    return {
        "mensaje": "Importación completada",
        "ingredientes": {
            "creados": len(resultado_ingredientes["creados"]),
            "actualizados": len(resultado_ingredientes["actualizados"]),
            "errores": len(resultado_ingredientes["errores"]),
            "detalle": resultado_ingredientes,
        },
        "recetas": {
            "creadas": len(resultado_recetas["creadas"]),
            "omitidas": len(resultado_recetas["omitidas"]),
            "errores": len(resultado_recetas["errores"]),
            "detalle": resultado_recetas,
        },
    }
