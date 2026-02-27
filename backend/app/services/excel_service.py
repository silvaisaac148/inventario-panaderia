"""Servicio: Excel — importar y exportar datos."""

import io
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from app.models.ingrediente import Ingrediente
from app.models.producto import Producto
from app.models.movimiento import MovimientoStock


# ============================================================
#  IMPORTACIÓN
# ============================================================

async def importar_ingredientes_excel(
    db: AsyncSession, file_bytes: bytes
) -> dict:
    """
    Importa ingredientes desde un archivo Excel.
    Formato esperado:
        nombre | unidad_medida | peso_compra | costo | stock_inicial | stock_minimo
    """
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True)
    ws = wb.active

    # Leer encabezados
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    headers_lower = [h.lower().strip() if h else "" for h in headers]

    # Mapear columnas
    col_map = {}
    expected = ["nombre", "unidad_medida", "peso_compra", "costo", "stock_inicial", "stock_minimo"]
    for exp in expected:
        for i, h in enumerate(headers_lower):
            if exp in h or h in exp:
                col_map[exp] = i
                break

    if "nombre" not in col_map:
        raise ValueError(
            "El archivo debe tener al menos la columna 'nombre'. "
            f"Columnas encontradas: {headers}"
        )

    importados = 0
    errores = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            nombre = str(row[col_map.get("nombre", 0)] or "").strip()
            if not nombre:
                continue

            unidad = str(row[col_map.get("unidad_medida", 1)] or "kg").strip().lower()
            stock = float(row[col_map.get("stock_inicial", 4)] or 0)
            costo = float(row[col_map.get("costo", 3)] or 0)
            stock_min = float(row[col_map.get("stock_minimo", 5)] or 0)

            # Verificar si ya existe
            result = await db.execute(
                select(Ingrediente).where(Ingrediente.nombre == nombre)
            )
            existente = result.scalar_one_or_none()

            if existente:
                # Actualizar stock y costo
                existente.stock_actual = stock
                existente.costo_unitario = costo
                existente.stock_minimo = stock_min
            else:
                ingrediente = Ingrediente(
                    nombre=nombre,
                    unidad_medida=unidad if unidad in ("kg", "g", "lts", "ml", "und") else "kg",
                    stock_actual=stock,
                    stock_minimo=stock_min,
                    costo_unitario=costo,
                    costo_por_gramo=0,
                )
                db.add(ingrediente)

            importados += 1

        except Exception as e:
            errores.append({"fila": row_idx, "error": str(e)})

    wb.close()

    return {
        "importados": importados,
        "errores": errores,
        "total_filas": row_idx - 1 if row_idx else 0,
    }


async def importar_recetas_excel(
    db: AsyncSession, file_bytes: bytes
) -> dict:
    """
    Importa recetas desde un archivo Excel.
    Formato esperado:
        producto | ingrediente | cantidad | unidad
    """
    from app.models.receta import Receta, RecetaDetalle

    wb = load_workbook(io.BytesIO(file_bytes), read_only=True)
    ws = wb.active

    importados = 0
    errores = []
    recetas_creadas = {}

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            producto_nombre = str(row[0] or "").strip()
            ingrediente_nombre = str(row[1] or "").strip()
            cantidad = float(row[2] or 0)
            unidad = str(row[3] or "g").strip().lower()

            if not producto_nombre or not ingrediente_nombre:
                continue

            # Buscar producto
            result = await db.execute(
                select(Producto).where(Producto.nombre == producto_nombre)
            )
            producto = result.scalar_one_or_none()
            if not producto:
                # Crear producto si no existe
                producto = Producto(nombre=producto_nombre)
                db.add(producto)
                await db.flush()

            # Buscar ingrediente
            result = await db.execute(
                select(Ingrediente).where(Ingrediente.nombre == ingrediente_nombre)
            )
            ingrediente = result.scalar_one_or_none()
            if not ingrediente:
                errores.append({
                    "fila": row_idx,
                    "error": f"Ingrediente '{ingrediente_nombre}' no encontrado"
                })
                continue

            # Crear o reutilizar receta
            if producto_nombre not in recetas_creadas:
                receta = Receta(
                    producto_id=producto.id,
                    nombre=f"Receta {producto_nombre}",
                    rendimiento=1,
                )
                db.add(receta)
                await db.flush()
                recetas_creadas[producto_nombre] = receta
            else:
                receta = recetas_creadas[producto_nombre]

            # Agregar detalle
            detalle = RecetaDetalle(
                receta_id=receta.id,
                ingrediente_id=ingrediente.id,
                cantidad=cantidad,
                unidad=unidad,
            )
            db.add(detalle)
            importados += 1

        except Exception as e:
            errores.append({"fila": row_idx, "error": str(e)})

    wb.close()

    return {
        "recetas_creadas": len(recetas_creadas),
        "detalles_importados": importados,
        "errores": errores,
    }


# ============================================================
#  EXPORTACIÓN
# ============================================================

def _estilo_encabezado():
    return {
        "font": Font(bold=True, color="FFFFFF", size=11),
        "fill": PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid"),
        "alignment": Alignment(horizontal="center", vertical="center"),
        "border": Border(
            bottom=Side(style="thin"),
            top=Side(style="thin"),
            left=Side(style="thin"),
            right=Side(style="thin"),
        ),
    }


async def exportar_stock_excel(db: AsyncSession) -> bytes:
    """Exporta el stock actual de ingredientes a un archivo Excel."""
    result = await db.execute(
        select(Ingrediente).order_by(Ingrediente.nombre)
    )
    ingredientes = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Ingredientes"

    # Encabezados
    headers = ["Nombre", "Unidad", "Stock Actual", "Stock Mínimo", "Costo Unitario", "Alerta"]
    estilos = _estilo_encabezado()
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = estilos["font"]
        cell.fill = estilos["fill"]
        cell.alignment = estilos["alignment"]
        cell.border = estilos["border"]

    # Datos
    for row_idx, ing in enumerate(ingredientes, 2):
        ws.cell(row=row_idx, column=1, value=ing.nombre)
        ws.cell(row=row_idx, column=2, value=ing.unidad_medida)
        ws.cell(row=row_idx, column=3, value=float(ing.stock_actual))
        ws.cell(row=row_idx, column=4, value=float(ing.stock_minimo))
        ws.cell(row=row_idx, column=5, value=float(ing.costo_unitario))

        alerta = "⚠️ BAJO" if float(ing.stock_actual) <= float(ing.stock_minimo) else "✅ OK"
        cell = ws.cell(row=row_idx, column=6, value=alerta)
        if alerta.startswith("⚠️"):
            cell.fill = PatternFill(start_color="FFE0E0", fill_type="solid")

    # Ajustar ancho de columnas
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 18

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


async def exportar_movimientos_excel(db: AsyncSession) -> bytes:
    """Exporta el historial de movimientos de stock."""
    from sqlalchemy.orm import selectinload

    result = await db.execute(
        select(MovimientoStock)
        .options(selectinload(MovimientoStock.ingrediente))
        .order_by(MovimientoStock.fecha.desc())
    )
    movimientos = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Movimientos"

    headers = ["Fecha", "Ingrediente", "Tipo", "Cantidad", "Costo Total", "Referencia"]
    estilos = _estilo_encabezado()
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = estilos["font"]
        cell.fill = estilos["fill"]
        cell.alignment = estilos["alignment"]

    for row_idx, mov in enumerate(movimientos, 2):
        ws.cell(row=row_idx, column=1, value=mov.fecha.strftime("%Y-%m-%d %H:%M"))
        ws.cell(row=row_idx, column=2, value=mov.ingrediente.nombre if mov.ingrediente else "")
        ws.cell(row=row_idx, column=3, value=mov.tipo)
        ws.cell(row=row_idx, column=4, value=float(mov.cantidad))
        ws.cell(row=row_idx, column=5, value=float(mov.costo_total or 0))
        ws.cell(row=row_idx, column=6, value=mov.referencia or "")

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 20

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
